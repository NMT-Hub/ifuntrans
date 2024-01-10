import argparse
import asyncio
from typing import List

import langcodes
import openpyxl
import pandas as pd
from loguru import logger
from openpyxl.styles import Font

from ifuntrans.api.localization import normalize_case
from ifuntrans.async_translators.chatgpt import normalize_language_code_as_iso639
from ifuntrans.tm import create_tm_from_excel
from ifuntrans.translate import translate


async def translate_group(
    dataframe: pd.DataFrame,
    langcodes_2_columns: dict,
    columns_2_langcodes: dict,
    instructions: str = "",
    tm=None,
):
    # find the most common language code. more specifically, "zh" and "en"
    # zh and en will translate to each other
    # the others will be translated from these two languages
    try:
        zh_lang_code = langcodes.closest_supported_match("zh", langcodes_2_columns.keys())
        zh_column = langcodes_2_columns[zh_lang_code]
    except KeyError:
        dataframe["zh"] = pd.NA
        zh_column = "zh"
        columns_2_langcodes["zh"] = "zh"

    try:
        en_lang_code = langcodes.closest_supported_match("en", langcodes_2_columns.keys())
        en_column = langcodes_2_columns[en_lang_code]
    except KeyError:
        dataframe["en"] = pd.NA
        en_column = "en"
        columns_2_langcodes["en"] = "en"

    async def translate_column(
        source_column: str,
        target_column: str,
    ):
        from_lang = columns_2_langcodes[source_column]
        to_lang = columns_2_langcodes[target_column]

        # select rows that target_column is empty and source_column is not empty
        # these rows will be translated from 'from_lang' to 'to_lang'
        target_empty_rows = dataframe[dataframe[target_column].isnull() & dataframe[source_column].notnull()]
        # group by "source_column"
        # if there are multiple rows with the same "source_column", we translate them together
        # this will save the translation cost
        dedup_group = target_empty_rows.groupby(source_column)
        translation: List[str] = await translate(
            dedup_group.groups.keys(),
            from_lang=from_lang,
            to_lang=to_lang,
            tm=tm,
            instructions=instructions,
        )
        # update the group
        dataframe.loc[target_empty_rows.index, target_column] = dedup_group[target_column].transform(
            lambda _: translation.pop(0)
        )

    await translate_column(en_column, zh_column)
    await translate_column(zh_column, en_column)

    for column in columns_2_langcodes.keys():
        if column in [zh_column, en_column]:
            continue

        lang_code = columns_2_langcodes[column]

        if langcodes.get(lang_code).language in ["zh", "ja", "ko"]:
            await translate_column(zh_column, column)
        else:
            await translate_column(en_column, column)


async def main():
    parser = argparse.ArgumentParser(description="Translate excel file")
    parser.add_argument("file", help="The excel file to translate")
    parser.add_argument("--output", help="The output file", default=None, type=str)
    parser.add_argument("-l", "--language", nargs="+", help="The additional languages to translate to", default=[])
    parser.add_argument("-tm", "--translate-memory-file", help="The translate memory file", default=None, type=str)
    parser.add_argument("-i", "--instructions", help="The instructions prompt", default="", type=str)
    parser.add_argument("-k", "--keep-format", help="Keep the format of the original file", action="store_true")

    args = parser.parse_args()
    if args.translate_memory_file is not None:
        await create_tm_from_excel(args.translate_memory_file)
    else:
        pass

    if args.output is None:
        args.output = args.file

    workbook = openpyxl.load_workbook(args.file)
    sheet_names = workbook.sheetnames

    sheet_2_dataframes = {}
    for sheet_name in sheet_names:
        # incase "None" replace by pd.NA
        dataframe = pd.read_excel(args.file, sheet_name=sheet_name, keep_default_na=False)
        # convert all columns to string
        dataframe = dataframe.astype(str)
        # apply str.strip to all columns
        dataframe = dataframe.apply(lambda x: x.str.strip())
        # apply normalize_case to all columns
        dataframe = dataframe.applymap(normalize_case)
        # manually replace "" with pd.NA
        dataframe.replace("", pd.NA, inplace=True)
        dataframe.replace("nan", pd.NA, inplace=True)
        columns = dataframe.columns

        # normalize the language code
        iso_codes = await normalize_language_code_as_iso639(columns[1:])
        iso_codes.insert(0, "und")  # the first column is the key

        if not langcodes.closest_supported_match("zh", iso_codes) and not langcodes.closest_supported_match(
            "en", iso_codes
        ):
            logger.warning(f"Sheet {sheet_name} has no column with valid language code")
            continue

        if args.language:
            for i, code in enumerate(iso_codes):
                if not langcodes.closest_supported_match(code, args.language + ["zh", "en"]):
                    iso_codes[i] = "und"

        columns_2_langcodes = {}
        for column, code in zip(columns, iso_codes):
            if code == "und":
                continue
            if code not in columns_2_langcodes.values():
                columns_2_langcodes[column] = code
            else:
                langcodes_2_columns = {v: k for k, v in columns_2_langcodes.items()}
                await translate_group(dataframe, langcodes_2_columns, columns_2_langcodes, args.instructions)
                columns_2_langcodes = {column: code}

        if len(columns_2_langcodes) > 0:
            langcodes_2_columns = {v: k for k, v in columns_2_langcodes.items()}
            await translate_group(dataframe, langcodes_2_columns, columns_2_langcodes, args.instructions)

        sheet_2_dataframes[sheet_name] = dataframe

    if args.keep_format:
        for sheet_name, dataframe in sheet_2_dataframes.items():
            ws = workbook[sheet_name]
            dataframe = sheet_2_dataframes[sheet_name]
            # Update the localization workbook
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value is None:
                        try:
                            cell.value = dataframe.iloc[cell.row - 2, cell.column - 1]
                        except (IndexError, ValueError, AttributeError):
                            continue
                        cell.font = Font(color="FF0000")
        workbook.save(args.output)
    else:
        with pd.ExcelWriter(args.output) as writer:
            for sheet_name, dataframe in sheet_2_dataframes.items():
                dataframe.to_excel(writer, sheet_name=sheet_name, index=False)


if __name__ == "__main__":
    asyncio.run(main())
